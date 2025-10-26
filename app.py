
import os
import json
import datetime
from pathlib import Path
from io import BytesIO
from flask import (
    Flask, request, redirect, url_for, send_from_directory, flash,
    jsonify, render_template_string, send_file
)
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

# =========================================================
# Inicijalizacija (NE DIRATI – buduće sekcije samo dodaj ispod)
# =========================================================
app = Flask(__name__)
app.secret_key = "promijeni-ovo-u-jak-tajni-kljuc"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads"
DOCS_DIR = UPLOAD_DIR / "docs"
IMG_DIR = UPLOAD_DIR / "images"
MEMBER_DIR = UPLOAD_DIR / "members"

for p in [DATA_DIR, UPLOAD_DIR, DOCS_DIR, IMG_DIR, MEMBER_DIR]:
    p.mkdir(parents=True, exist_ok=True)

CLUB_JSON = DATA_DIR / "club.json"
MEMBERS_JSON = DATA_DIR / "members.json"

ALLOWED_DOC_EXT = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".jpg", ".jpeg", ".png"}
ALLOWED_IMG_EXT = {".jpg", ".jpeg", ".png", ".svg", ".webp"}
ALLOWED_XLSX = {".xlsx"}

# =========================================================
# Helperi (NE DIRATI)
# =========================================================
def allowed_file(filename, allowed_exts):
    ext = os.path.splitext(filename)[1].lower()
    return ext in allowed_exts

def safe_save(file_storage, target_dir: Path, allowed_exts):
    if not file_storage or file_storage.filename.strip() == "":
        return ""
    filename = file_storage.filename
    if not allowed_file(filename, allowed_exts):
        raise ValueError("Nedozvoljen tip datoteke.")
    ext = os.path.splitext(filename)[1].lower()
    base = os.path.splitext(os.path.basename(filename))[0]
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{base}_{timestamp}{ext}"
    path = target_dir / safe_name
    file_storage.save(path)
    # vratimo public path
    sub = "images" if target_dir == IMG_DIR else ("members" if target_dir == MEMBER_DIR else "docs")
    return f"/uploads/{sub}/{safe_name}"

def load_json(path: Path, default):
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def save_json(path: Path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_club():
    return load_json(CLUB_JSON, {
        "naziv": "HRVAČKI KLUB PODRAVKA",
        "ulica_broj": "Miklinovec 6a",
        "grad_postanski": "48000 Koprivnica",
        "oib": "60911784858",
        "iban": "HR6923860021100518154",
        "email": "hsk-podravka@gmail.com",
        "web": "https://hk-podravka.com",
        "social": {"instagram": "", "facebook": "", "tiktok": ""},
        "logo_path": "",
        "predsjednik": {"ime_prezime": "", "telefon": "", "email": ""},
        "tajnik": {"ime_prezime": "", "telefon": "", "email": ""},
        "predsjednistvo": [],
        "nadzorni_odbor": [],
        "statut_path": "",
        "ostali_dokumenti": []
    })

def save_club(data): save_json(CLUB_JSON, data)

def load_members():
    return load_json(MEMBERS_JSON, {"seq": 1, "items": []})

def save_members(data): save_json(MEMBERS_JSON, data)

def new_member_id(members):
    mid = members["seq"]
    members["seq"] += 1
    return mid

def parse_date(s):
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except:
        return None

def days_to(date_val: datetime.date):
    if not date_val: return None
    return (date_val - datetime.date.today()).days

def as_bool(val):
    return str(val).lower() in {"1","true","on","da","yes"}

# =========================================================
# Statika za upload-e (NE DIRATI)
# =========================================================
@app.route("/uploads/<path:filename>")
def uploaded_files(filename):
    if filename.startswith("images/"):
        return send_from_directory(IMG_DIR, filename.replace("images/", ""))
    if filename.startswith("members/"):
        return send_from_directory(MEMBER_DIR, filename.replace("members/", ""))
    return send_from_directory(DOCS_DIR, filename.replace("docs/", ""))

# =========================================================
# SEKCIJA 1 — OSNOVNI PODACI O KLUBU (NE DIRATI)
# =========================================================
@app.route("/")
def home_redirect():
    return redirect(url_for("klub"))

@app.route("/klub", methods=["GET", "POST"])
def klub():
    club = load_club()
    if request.method == "POST":
        try:
            club["naziv"] = request.form.get("naziv", "").strip()
            club["ulica_broj"] = request.form.get("ulica_broj", "").strip()
            club["grad_postanski"] = request.form.get("grad_postanski", "").strip()
            club["oib"] = request.form.get("oib", "").strip()
            club["iban"] = request.form.get("iban", "").strip()
            club["email"] = request.form.get("email", "").strip()
            club["web"] = request.form.get("web", "").strip()

            club["social"]["instagram"] = request.form.get("instagram", "").strip()
            club["social"]["facebook"] = request.form.get("facebook", "").strip()
            club["social"]["tiktok"] = request.form.get("tiktok", "").strip()

            club["predsjednik"] = {
                "ime_prezime": request.form.get("pred_ime_prezime", "").strip(),
                "telefon": request.form.get("pred_telefon", "").strip(),
                "email": request.form.get("pred_email", "").strip(),
            }
            club["tajnik"] = {
                "ime_prezime": request.form.get("taj_ime_prezime", "").strip(),
                "telefon": request.form.get("taj_telefon", "").strip(),
                "email": request.form.get("taj_email", "").strip(),
            }

            # dinamika
            p_count = int(request.form.get("predsjednistvo_count", "0"))
            predsjednistvo = []
            for i in range(p_count):
                ime = request.form.get(f"p_{i}_ime","").strip()
                tel = request.form.get(f"p_{i}_telefon","").strip()
                eml = request.form.get(f"p_{i}_email","").strip()
                if any([ime,tel,eml]):
                    predsjednistvo.append({"ime_prezime": ime, "telefon": tel, "email": eml})
            club["predsjednistvo"] = predsjednistvo

            n_count = int(request.form.get("nadzor_count", "0"))
            nadzor = []
            for i in range(n_count):
                ime = request.form.get(f"n_{i}_ime","").strip()
                tel = request.form.get(f"n_{i}_telefon","").strip()
                eml = request.form.get(f"n_{i}_email","").strip()
                if any([ime,tel,eml]):
                    nadzor.append({"ime_prezime": ime, "telefon": tel, "email": eml})
            club["nadzorni_odbor"] = nadzor

            if "logo" in request.files and request.files["logo"].filename:
                club["logo_path"] = safe_save(request.files["logo"], IMG_DIR, ALLOWED_IMG_EXT)

            if "statut" in request.files and request.files["statut"].filename:
                club["statut_path"] = safe_save(request.files["statut"], DOCS_DIR, ALLOWED_DOC_EXT)

            if "ostali_dokumenti" in request.files:
                lst = club.get("ostali_dokumenti", []) or []
                for f in request.files.getlist("ostali_dokumenti"):
                    if f and f.filename:
                        try:
                            lst.append(safe_save(f, DOCS_DIR, ALLOWED_DOC_EXT))
                        except ValueError:
                            pass
                club["ostali_dokumenti"] = lst

            save_club(club)
            flash("Osnovni podaci spremljeni.", "success")
            return redirect(url_for("klub"))
        except Exception as e:
            flash(f"Greška: {e}", "danger")
    return render_template_string(TPL_KLUB, club=club)

@app.route("/api/club")
def api_club(): return jsonify(load_club())

# =========================================================
# SEKCIJA 2 — ČLANOVI (NE DIRATI)
# =========================================================
MEMBER_FIELDS = [
    "ime_prezime","datum_rodjenja","spol","oib","mjesto",
    "email_sportas","email_roditelj",
    "osobna_broj","osobna_vrijedi_do","osobna_izdao",
    "putovnica_broj","putovnica_vrijedi_do","putovnica_izdao",
    "aktivan_natjecatelj","veteran","ostalo",
    "placa_clanarinu","iznos_clanarine","grupa",
]

@app.route("/clanovi", methods=["GET","POST"])
def clanovi():
    members = load_members()
    club = load_club()

    if request.method == "POST":
        form = request.form
        files = request.files

        # novo ili update?
        mid = form.get("member_id")
        is_update = bool(mid)

        def get_num(field, default=0.0):
            try: return float(form.get(field, default))
            except: return default

        member = {
            "id": int(mid) if is_update else new_member_id(members),
            "ime_prezime": form.get("ime_prezime","").strip(),
            "datum_rodjenja": form.get("datum_rodjenja","").strip(),
            "spol": form.get("spol","").strip(),
            "oib": form.get("oib","").strip(),
            "mjesto": form.get("mjesto","").strip(),
            "email_sportas": form.get("email_sportas","").strip(),
            "email_roditelj": form.get("email_roditelj","").strip(),
            "osobna_broj": form.get("osobna_broj","").strip(),
            "osobna_vrijedi_do": form.get("osobna_vrijedi_do","").strip(),
            "osobna_izdao": form.get("osobna_izdao","").strip(),
            "putovnica_broj": form.get("putovnica_broj","").strip(),
            "putovnica_vrijedi_do": form.get("putovnica_vrijedi_do","").strip(),
            "putovnica_izdao": form.get("putovnica_izdao","").strip(),
            "aktivan_natjecatelj": as_bool(form.get("aktivan_natjecatelj")),
            "veteran": as_bool(form.get("veteran")),
            "ostalo": as_bool(form.get("ostalo")),
            "placa_clanarinu": as_bool(form.get("placa_clanarinu")),
            "iznos_clanarine": get_num("iznos_clanarine", 30.0),
            "grupa": form.get("grupa","").strip(),
            "slika_path": "",
            "pristupnica_pdf": "",
            "privola_pdf": "",
            "pristupnica_upload": "",
            "privola_upload": "",
            "lijecnicka_upload": "",
            "lijecnicka_vrijedi_do": form.get("lijecnicka_vrijedi_do","").strip(),
            "datum_pristupa": form.get("datum_pristupa","").strip() or datetime.date.today().isoformat(),
        }

        # slika
        if "slika" in files and files["slika"].filename:
            member["slika_path"] = safe_save(files["slika"], MEMBER_DIR, ALLOWED_IMG_EXT)

        # dokumenti upload
        if "pristupnica_upload" in files and files["pristupnica_upload"].filename:
            member["pristupnica_upload"] = safe_save(files["pristupnica_upload"], MEMBER_DIR, ALLOWED_DOC_EXT)
        if "privola_upload" in files and files["privola_upload"].filename:
            member["privola_upload"] = safe_save(files["privola_upload"], MEMBER_DIR, ALLOWED_DOC_EXT)
        if "lijecnicka_upload" in files and files["lijecnicka_upload"].filename:
            member["lijecnicka_upload"] = safe_save(files["lijecnicka_upload"], MEMBER_DIR, ALLOWED_DOC_EXT)

        # Ako update – zamijeni unesenog
        if is_update:
            members["items"] = [m for m in members["items"] if int(m["id"]) != int(mid)]
        members["items"].append(member)
        save_members(members)

        # Auto generiranje PDF-ova (pristupnica + privola)
        gen_pristupnica_pdf(member, club)
        gen_privola_pdf(member, club)
        save_members(members)

        flash("Član je spremljen. PDF pristupnica i privola su generirani.", "success")
        return redirect(url_for("clanovi"))

    # GET
    items = sorted(load_members()["items"], key=lambda x: (x.get("ime_prezime",""), x["id"]))
    # statusi liječničke
    for it in items:
        it["lijecnicka_dana_preostalo"] = days_to(parse_date(it.get("lijecnicka_vrijedi_do")))
    return render_template_string(TPL_CLANOVI, club=club, members=items)

@app.route("/clanovi/obrisi/<int:mid>", methods=["POST"])
def clanovi_obrisi(mid):
    members = load_members()
    before = len(members["items"])
    members["items"] = [m for m in members["items"] if int(m["id"]) != int(mid)]
    save_members(members)
    flash("Član obrisan." if len(members["items"])<before else "Član nije pronađen.", "info")
    return redirect(url_for("clanovi"))

@app.route("/clanovi/predlozak")
def clanovi_predlozak():
    # Excel predložak s točnim kolonama redoslijedom
    wb = Workbook()
    ws = wb.active
    ws.title = "Clanovi"
    headers = [
        "ime_prezime","datum_rodjenja(YYYY-MM-DD)","spol(m/ž)","oib","mjesto",
        "email_sportas","email_roditelj",
        "osobna_broj","osobna_vrijedi_do(YYYY-MM-DD)","osobna_izdao",
        "putovnica_broj","putovnica_vrijedi_do(YYYY-MM-DD)","putovnica_izdao",
        "aktivan_natjecatelj(da/ne)","veteran(da/ne)","ostalo(da/ne)",
        "placa_clanarinu(da/ne)","iznos_clanarine(decimal)","grupa",
        "lijecnicka_vrijedi_do(YYYY-MM-DD)"
    ]
    ws.append(headers)
    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="predlozak_clanovi.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/clanovi/izvoz")
def clanovi_izvoz():
    # izvoz postojeće baze u xlsx
    data = load_members()["items"]
    wb = Workbook(); ws = wb.active; ws.title = "Clanovi"
    headers = [
        "id","ime_prezime","datum_rodjenja","spol","oib","mjesto",
        "email_sportas","email_roditelj",
        "osobna_broj","osobna_vrijedi_do","osobna_izdao",
        "putovnica_broj","putovnica_vrijedi_do","putovnica_izdao",
        "aktivan_natjecatelj","veteran","ostalo",
        "placa_clanarinu","iznos_clanarine","grupa",
        "datum_pristupa","lijecnicka_vrijedi_do"
    ]
    ws.append(headers)
    for m in data:
        ws.append([
            m.get("id"), m.get("ime_prezime"), m.get("datum_rodjenja"), m.get("spol"),
            m.get("oib"), m.get("mjesto"), m.get("email_sportas"), m.get("email_roditelj"),
            m.get("osobna_broj"), m.get("osobna_vrijedi_do"), m.get("osobna_izdao"),
            m.get("putovnica_broj"), m.get("putovnica_vrijedi_do"), m.get("putovnica_izdao"),
            "da" if m.get("aktivan_natjecatelj") else "ne",
            "da" if m.get("veteran") else "ne",
            "da" if m.get("ostalo") else "ne",
            "da" if m.get("placa_clanarinu") else "ne",
            m.get("iznos_clanarine"),
            m.get("grupa"),
            m.get("datum_pristupa"),
            m.get("lijecnicka_vrijedi_do"),
        ])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="clanovi_izvoz.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/clanovi/upload", methods=["POST"])
def clanovi_upload():
    members = load_members()
    f = request.files.get("excel")
    if not f or f.filename == "":
        flash("Nije odabran Excel.", "warning")
        return redirect(url_for("clanovi"))
    if not allowed_file(f.filename, ALLOWED_XLSX):
        flash("Dozvoljen je samo .xlsx", "danger")
        return redirect(url_for("clanovi"))

    bio = BytesIO(f.read()); bio.seek(0)
    wb = load_workbook(bio); ws = wb.active
    # očekujemo zaglavlje kao u predlošku
    rows = list(ws.iter_rows(values_only=True))
    if not rows: 
        flash("Prazan Excel.", "warning"); 
        return redirect(url_for("clanovi"))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # minimalna provjera
    if "ime_prezime" not in header:
        flash("Predložak nije ispravan (nedostaje 'ime_prezime').", "danger")
        return redirect(url_for("clanovi"))

    idx = {h:i for i,h in enumerate(header)}
    def val(row, key):
        i = idx.get(key, None)
        return (str(row[i]).strip() if i is not None and row[i] is not None else "")

    count = 0
    for row in rows[1:]:
        if not row: continue
        name = val(row,"ime_prezime")
        if not name: continue
        m = {
            "id": new_member_id(members),
            "ime_prezime": name,
            "datum_rodjenja": val(row,"datum_rodjenja(YYYY-MM-DD)") or val(row,"datum_rodjenja"),
            "spol": val(row,"spol(m/ž)") or val(row,"spol"),
            "oib": val(row,"oib"),
            "mjesto": val(row,"mjesto"),
            "email_sportas": val(row,"email_sportas"),
            "email_roditelj": val(row,"email_roditelj"),
            "osobna_broj": val(row,"osobna_broj"),
            "osobna_vrijedi_do": val(row,"osobna_vrijedi_do(YYYY-MM-DD)") or val(row,"osobna_vrijedi_do"),
            "osobna_izdao": val(row,"osobna_izdao"),
            "putovnica_broj": val(row,"putovnica_broj"),
            "putovnica_vrijedi_do": val(row,"putovnica_vrijedi_do(YYYY-MM-DD)") or val(row,"putovnica_vrijedi_do"),
            "putovnica_izdao": val(row,"putovnica_izdao"),
            "aktivan_natjecatelj": (val(row,"aktivan_natjecatelj(da/ne)") or val(row,"aktivan_natjecatelj")).lower()=="da",
            "veteran": (val(row,"veteran(da/ne)") or val(row,"veteran")).lower()=="da",
            "ostalo": (val(row,"ostalo(da/ne)") or val(row,"ostalo")).lower()=="da",
            "placa_clanarinu": (val(row,"placa_clanarinu(da/ne)") or val(row,"placa_clanarinu")).lower()=="da",
            "iznos_clanarine": float((val(row,"iznos_clanarine(decimal)") or val(row,"iznos_clanarine") or "30").replace(",",".")),
            "grupa": val(row,"grupa"),
            "slika_path": "",
            "pristupnica_pdf": "",
            "privola_pdf": "",
            "pristupnica_upload": "",
            "privola_upload": "",
            "lijecnicka_upload": "",
            "lijecnicka_vrijedi_do": val(row,"lijecnicka_vrijedi_do(YYYY-MM-DD)") or "",
            "datum_pristupa": datetime.date.today().isoformat(),
        }
        members["items"].append(m)
        count += 1
    save_members(members)
    flash(f"Uvezeno članova: {count}", "success")
    return redirect(url_for("clanovi"))

# -------- PDF generatori (NE DIRATI) --------
def gen_pristupnica_pdf(member, club):
    # generira i snima PDF te ažurira path u članu
    path = MEMBER_DIR / f"pristupnica_{member['id']}.pdf"
    c = canvas.Canvas(str(path), pagesize=A4)
    w, h = A4
    top = h - 20*mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20*mm, top, "PRISTUPNICA – HRVAČKI KLUB PODRAVKA")
    c.setFont("Helvetica", 11)
    c.drawString(20*mm, top-10*mm, f"Ime i prezime: {member.get('ime_prezime','')}")
    c.drawString(20*mm, top-16*mm, f"Datum rođenja: {member.get('datum_rodjenja','')}")
    c.drawString(20*mm, top-22*mm, f"OIB: {member.get('oib','')}")
    text = c.beginText(20*mm, top-35*mm)
    text.setFont("Helvetica", 10)
    # Skraćeni header s tvojim tekstom (ostatak teksta preuzet iz upita)
    paragraf = (
        "HRVAČKI KLUB „PODRAVKA“ 48000 Koprivnica, Miklinovec 6a, "
        "mob:091/456-23-21 web site: www.hk-podravka.hr, e-mail: hsk.podravka@gmail.com\n\n"
        "OIB:60911784858, žiro-račun: HR6923860021100518154, Podravska banka d.d. Koprivnica\n\n"
        "STATUT KLUBA - ČLANSTVO\n"
        "Članak 14. Članom Kluba može postati svaki poslovno sposoban državljanin RH... "
        "(Napomena: Cijeli Statut dostupan je na www.hk-podravka.hr/o-klubu)\n\n"
        "STATUT KLUBA – PRESTANAK ČLANSTVA\n"
        "Članak 21. Članstvo u klubu prestaje: dragovoljnim istupom... "
        "(Napomena: Istupnica je dostupna na www.hk-podravka.hr/o-klubu)\n\n"
        "ČLANARINA JE OBVEZUJUĆA TIJEKOM CIJELE GODINE (12 MJESECI)...\n\n"
        "IZJAVA O ODGOVORNOSTI\n"
        "Hrvanje je borilački sport... Svojim potpisom suglasni smo...\n\n"
        "POTPIS ČLANA: __________________________\n"
        "POTPIS RODITELJA/STARATELJA: ______________________\n"
    )
    for line in paragraf.split("\n"):
        text.textLine(line)
    c.drawText(text)
    c.showPage(); c.save()
    # spremi path
    members = load_members()
    for m in members["items"]:
        if m["id"] == member["id"]:
            m["pristupnica_pdf"] = f"/uploads/members/{path.name}"
            break
    save_members(members)

def gen_privola_pdf(member, club):
    path = MEMBER_DIR / f"privola_{member['id']}.pdf"
    c = canvas.Canvas(str(path), pagesize=A4)
    w, h = A4
    y = h - 20*mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20*mm, y, "PRIVOLA – obrada osobnih podataka (GDPR)")
    y -= 12*mm
    c.setFont("Helvetica", 10)
    text = c.beginText(20*mm, y)
    paragraf = (
        "Sukladno Zakonu o zaštiti osobnih podataka i Uredbi (EU) 2016/679 (GDPR)...\n\n"
        "OVIME DAJEM PRIVOLU da se moji osobni podaci (ime i prezime, OIB, datum rođenja, adresa, "
        "fotografija, video snimka, kontakt roditelja, e-mail, broj dokumenata, potvrda o zdravstvenoj sposobnosti, "
        "specifična medicinska stanja, ime škole/fakulteta...) koriste u svrhu vođenja i redovnog funkcioniranja Kluba, "
        "prijava i sudjelovanja na natjecanjima, te objava na službenim stranicama i društvenim mrežama Kluba i nadležnih tijela.\n\n"
        "U _____________________________ ; _____________ 20____\n"
        f"Član kluba: {member.get('ime_prezime','')}\n"
        "Potpis: ____________________\n"
        "Roditelj/staratelj malodobnog člana kluba: ________________________________\n"
        "Potpis roditelja/staratelja: ____________________\n"
    )
    for line in paragraf.split("\n"):
        text.textLine(line)
    c.drawText(text)
    c.showPage(); c.save()
    members = load_members()
    for m in members["items"]:
        if m["id"] == member["id"]:
            m["privola_pdf"] = f"/uploads/members/{path.name}"
            break
    save_members(members)

# =========================================================
# HTML TEMPLATES (NE DIRATI)
# =========================================================
TPL_BASE_HEAD = r"""
<!doctype html>
<html lang="hr">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>HK Podravka – Admin</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    :root{ --primary:#d90429; --accent:#c9a227; --light:#ffffff; }
    body{ background:#fafafa; }
    .navbar{ background:var(--primary); }
    .navbar .navbar-brand,.navbar .nav-link,.navbar .navbar-text{ color:var(--light)!important; }
    .card{ border-radius:1rem; box-shadow:0 8px 24px rgba(0,0,0,.06); }
    .btn-primary{ background:var(--primary); border-color:var(--primary); }
    .btn-outline-accent{ border-color:var(--accent); color:var(--accent); }
    .btn-outline-accent:hover{ background:var(--accent); color:var(--light); }
    .chip{ display:inline-block; padding:.25rem .5rem; border-radius:999px; border:1px solid #ddd; margin-right:.5rem; font-size:.85rem;}
    .section-title{ border-left:6px solid var(--accent); padding-left:.75rem; margin-top:1.5rem; margin-bottom:.75rem; }
    .logo-preview{ max-height:80px; object-fit:contain; }
    .danger-badge{ background:#ffebee; color:#b00020; border:1px solid #ffcdd2; padding:.15rem .5rem; border-radius:8px;}
  </style>
</head>
<body>
  <nav class="navbar navbar-expand-lg">
    <div class="container">
      <a class="navbar-brand fw-bold" href="{{ url_for('klub') }}">HK Podravka – Administracija</a>
      <div class="ms-auto d-flex gap-3">
        <a class="nav-link" href="{{ url_for('klub') }}">Klub</a>
        <a class="nav-link" href="{{ url_for('clanovi') }}">Članovi</a>
      </div>
    </div>
  </nav>
  <div class="container my-4">
    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        <div class="mb-3">
          {% for category, message in messages %}
            <div class="alert alert-{{category}} mb-2">{{message}}</div>
          {% endfor %}
        </div>
      {% endif %}
    {% endwith %}
"""

TPL_BASE_FOOT = r"""
  </div>
  <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
"""

TPL_KLUB = TPL_BASE_HEAD + r"""
<form method="POST" enctype="multipart/form-data" class="row g-3">
  <div class="col-12">
    <div class="card p-3 p-md-4">
      <div class="d-flex align-items-center gap-3">
        <img src="{{ club.logo_path if club.logo_path else 'https://dummyimage.com/280x120/ffffff/000000&text=LOGO' }}" class="logo-preview border rounded p-1 bg-white" alt="Logo">
        <div>
          <h1 class="h4 mb-0">{{ club.naziv }}</h1>
          <div class="text-muted small">
            {{ club.ulica_broj }}, {{ club.grad_postanski }}
          </div>
          <div class="mt-1">
            <span class="chip">OIB: {{ club.oib }}</span>
            <span class="chip">IBAN: {{ club.iban }}</span>
          </div>
        </div>
      </div>
      <hr>
      <div class="row g-3">
        <div class="col-md-6">
          <label class="form-label">Naziv kluba</label>
          <input name="naziv" class="form-control" value="{{ club.naziv }}">
        </div>
        <div class="col-md-6">
          <label class="form-label">Logo (jpg, png, svg, webp)</label>
          <input type="file" name="logo" accept=".jpg,.jpeg,.png,.svg,.webp" class="form-control">
        </div>
        <div class="col-md-6">
          <label class="form-label">Ulica i kućni broj</label>
          <input name="ulica_broj" class="form-control" value="{{ club.ulica_broj }}">
        </div>
        <div class="col-md-6">
          <label class="form-label">Grad i poštanski broj</label>
          <input name="grad_postanski" class="form-control" value="{{ club.grad_postanski }}">
        </div>
        <div class="col-md-4">
          <label class="form-label">OIB</label>
          <input name="oib" class="form-control" value="{{ club.oib }}">
        </div>
        <div class="col-md-8">
          <label class="form-label">IBAN</label>
          <input name="iban" class="form-control" value="{{ club.iban }}">
        </div>
        <div class="col-md-6">
          <label class="form-label">E-mail</label>
          <input name="email" type="email" class="form-control" value="{{ club.email }}">
        </div>
        <div class="col-md-6">
          <label class="form-label">Web stranica</label>
          <input name="web" type="url" class="form-control" value="{{ club.web }}">
        </div>
      </div>

      <h2 class="section-title h5">Društvene mreže</h2>
      <div class="row g-3">
        <div class="col-md-4"><label class="form-label">Instagram</label><input name="instagram" type="url" class="form-control" value="{{ club.social.instagram }}"></div>
        <div class="col-md-4"><label class="form-label">Facebook</label><input name="facebook" type="url" class="form-control" value="{{ club.social.facebook }}"></div>
        <div class="col-md-4"><label class="form-label">TikTok</label><input name="tiktok" type="url" class="form-control" value="{{ club.social.tiktok }}"></div>
      </div>

      <h2 class="section-title h5">Vodstvo kluba</h2>
      <div class="row g-3">
        <div class="col-md-4"><label class="form-label">Predsjednik – Ime i prezime</label><input name="pred_ime_prezime" class="form-control" value="{{ club.predsjednik.ime_prezime }}"></div>
        <div class="col-md-4"><label class="form-label">Predsjednik – Kontakt broj</label><input name="pred_telefon" class="form-control" value="{{ club.predsjednik.telefon }}"></div>
        <div class="col-md-4"><label class="form-label">Predsjednik – E-mail</label><input name="pred_email" type="email" class="form-control" value="{{ club.predsjednik.email }}"></div>
        <div class="col-md-4"><label class="form-label">Tajnik – Ime i prezime</label><input name="taj_ime_prezime" class="form-control" value="{{ club.tajnik.ime_prezime }}"></div>
        <div class="col-md-4"><label class="form-label">Tajnik – Kontakt broj</label><input name="taj_telefon" class="form-control" value="{{ club.tajnik.telefon }}"></div>
        <div class="col-md-4"><label class="form-label">Tajnik – E-mail</label><input name="taj_email" type="email" class="form-control" value="{{ club.tajnik.email }}"></div>
      </div>

      <h3 class="mt-4 h6">Članovi predsjedništva</h3>
      <div id="predsjednistvo-list" class="row g-3"></div>
      <button type="button" class="btn btn-outline-accent mt-1" onclick="addPresjednistvo()">+ Dodaj člana predsjedništva</button>
      <input type="hidden" id="predsjednistvo_count" name="predsjednistvo_count" value="0"/>

      <h3 class="mt-4 h6">Nadzorni odbor</h3>
      <div id="nadzor-list" class="row g-3"></div>
      <button type="button" class="btn btn-outline-accent mt-1" onclick="addNadzor()">+ Dodaj člana nadzornog odbora</button>
      <input type="hidden" id="nadzor_count" name="nadzor_count" value="0"/>

      <h2 class="section-title h5">Dokumenti</h2>
      <div class="row g-3">
        <div class="col-md-6">
          <label class="form-label">Statut (PDF)</label>
          <input type="file" name="statut" accept=".pdf" class="form-control">
          {% if club.statut_path %}
            <small class="d-block mt-1">Trenutni: <a href="{{ club.statut_path }}" target="_blank">{{ club.statut_path }}</a></small>
          {% endif %}
        </div>
        <div class="col-md-6">
          <label class="form-label">Ostali dokumenti</label>
          <input type="file" name="ostali_dokumenti" multiple class="form-control" accept=".pdf,.doc,.docx,.xls,.xlsx,.jpg,.jpeg,.png">
          {% if club.ostali_dokumenti %}
            <div class="mt-2">
              {% for d in club.ostali_dokumenti %}
                <div><a href="{{ d }}" target="_blank">{{ d }}</a></div>
              {% endfor %}
            </div>
          {% endif %}
        </div>
      </div>

      <div class="d-flex gap-2 mt-4">
        <button type="submit" class="btn btn-primary">Spremi</button>
        <a href="{{ url_for('api_club') }}" class="btn btn-outline-secondary">Pogledaj JSON</a>
      </div>
    </div>
  </div>
</form>

<script>
  const presetPredsjednistvo = {{ club.predsjednistvo|tojson }};
  const presetNadzor = {{ club.nadzorni_odbor|tojson }};
  let pCount = 0, nCount = 0;
  function memberRow(prefix, idx, data){
    const d = data || {ime_prezime:"", telefon:"", email:""};
    return `
      <div class="col-12">
        <div class="row g-2 align-items-end">
          <div class="col-md-5"><label class="form-label">Ime i prezime</label><input class="form-control" name="${prefix}_${idx}_ime" value="${d.ime_prezime||""}"></div>
          <div class="col-md-3"><label class="form-label">Kontakt broj</label><input class="form-control" name="${prefix}_${idx}_telefon" value="${d.telefon||""}"></div>
          <div class="col-md-4"><label class="form-label">E-mail</label><input type="email" class="form-control" name="${prefix}_${idx}_email" value="${d.email||""}"></div>
        </div>
      </div>`;
  }
  function addPresjednistvo(data){ const w=document.getElementById('predsjednistvo-list'); w.insertAdjacentHTML('beforeend', memberRow('p', pCount, data)); pCount++; document.getElementById('predsjednistvo_count').value=pCount; }
  function addNadzor(data){ const w=document.getElementById('nadzor-list'); w.insertAdjacentHTML('beforeend', memberRow('n', nCount, data)); nCount++; document.getElementById('nadzor_count').value=nCount; }
  window.addEventListener('DOMContentLoaded', ()=>{
    (presetPredsjednistvo.length?presetPredsjednistvo:[{}]).forEach(obj=>addPresjednistvo(obj));
    (presetNadzor.length?presetNadzor:[{}]).forEach(obj=>addNadzor(obj));
  });
</script>
""" + TPL_BASE_FOOT

TPL_CLANOVI = TPL_BASE_HEAD + r"""
<div class="card p-3 p-md-4 mb-4">
  <h1 class="h4 mb-3">Novi/uredi člana</h1>
  <form method="POST" enctype="multipart/form-data" class="row g-3">
    <input type="hidden" name="member_id" id="member_id">
    <div class="col-md-4"><label class="form-label">Ime i prezime</label><input name="ime_prezime" id="f_ime" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Datum rođenja</label><input type="date" name="datum_rodjenja" id="f_dat" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Spol</label>
      <select name="spol" id="f_spol" class="form-select">
        <option value="">—</option><option value="m">m</option><option value="ž">ž</option>
      </select></div>
    <div class="col-md-4"><label class="form-label">OIB</label><input name="oib" id="f_oib" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Mjesto prebivališta</label><input name="mjesto" id="f_mjesto" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Fotografija</label><input type="file" name="slika" accept=".jpg,.jpeg,.png,.svg,.webp" class="form-control"></div>

    <div class="col-md-4"><label class="form-label">E-mail sportaša</label><input type="email" name="email_sportas" id="f_em1" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">E-mail roditelja</label><input type="email" name="email_roditelj" id="f_em2" class="form-control"></div>

    <div class="col-md-4"><label class="form-label">Osobna iskaznica – broj</label><input name="osobna_broj" id="f_obr" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Osobna vrijedi do</label><input type="date" name="osobna_vrijedi_do" id="f_ovd" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Osobnu izdao</label><input name="osobna_izdao" id="f_oiz" class="form-control"></div>

    <div class="col-md-4"><label class="form-label">Putovnica – broj</label><input name="putovnica_broj" id="f_pbr" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Putovnica vrijedi do</label><input type="date" name="putovnica_vrijedi_do" id="f_pvd" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Putovnicu izdao</label><input name="putovnica_izdao" id="f_piz" class="form-control"></div>

    <div class="col-md-4"><label class="form-label">Grupa</label>
      <select name="grupa" id="f_grp" class="form-select">
        <option value="">—</option>
        <option value="Hrvači">Hrvači</option>
        <option value="Hrvačice">Hrvačice</option>
        <option value="Veterani">Veterani</option>
        <option value="Ostalo">Ostalo</option>
      </select>
    </div>

    <div class="col-md-8 d-flex align-items-center gap-3">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="aktivan_natjecatelj" id="f_act">
        <label class="form-check-label" for="f_act">Aktivni natjecatelj/ica</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="veteran" id="f_vet">
        <label class="form-check-label" for="f_vet">Veteran</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="ostalo" id="f_ost">
        <label class="form-check-label" for="f_ost">Ostalo</label>
      </div>
      <div class="form-check ms-4">
        <input class="form-check-input" type="checkbox" name="placa_clanarinu" id="f_fee" checked>
        <label class="form-check-label" for="f_fee">Plaća članarinu</label>
      </div>
      <div class="d-flex align-items-center gap-2">
        <span>Iznos (€):</span>
        <input name="iznos_clanarine" id="f_amt" class="form-control" style="width:120px" value="30">
      </div>
    </div>

    <div class="col-md-4"><label class="form-label">Datum pristupa (automatski)</label>
      <input name="datum_pristupa" id="f_dpr" type="date" class="form-control" value="{{ '%Y-%m-%d'|format }}">
    </div>

    <div class="col-md-4"><label class="form-label">Liječnička vrijedi do</label><input type="date" name="lijecnicka_vrijedi_do" class="form-control"></div>
    <div class="col-md-4"><label class="form-label">Liječnička potvrda (upload)</label><input type="file" name="lijecnicka_upload" class="form-control" accept=".pdf,.jpg,.jpeg,.png"></div>

    <div class="col-md-4"><label class="form-label">Pristupnica (upload)</label><input type="file" name="pristupnica_upload" class="form-control" accept=".pdf,.jpg,.jpeg,.png"></div>
    <div class="col-md-4"><label class="form-label">Privola (upload)</label><input type="file" name="privola_upload" class="form-control" accept=".pdf,.jpg,.jpeg,.png"></div>

    <div class="col-12 d-flex gap-2">
      <button class="btn btn-primary">Spremi člana</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('clanovi_predlozak') }}">Preuzmi Excel predložak</a>
      <a class="btn btn-outline-secondary" href="{{ url_for('clanovi_izvoz') }}">Izvoz članova (XLSX)</a>
    </div>

    <div class="col-12">
      <div class="alert alert-info small mb-0">
        * Za upis člana nije potrebno ispuniti sva polja — možete nadopuniti kasnije. PDF pristupnica i privola generiraju se automatski.
      </div>
    </div>
  </form>
  <hr class="my-4">
  <form action="{{ url_for('clanovi_upload') }}" method="POST" enctype="multipart/form-data" class="d-flex gap-2">
    <input class="form-control" type="file" name="excel" accept=".xlsx">
    <button class="btn btn-outline-accent">Uvezi članove (Excel .xlsx)</button>
  </form>
</div>

<div class="card p-3 p-md-4">
  <h2 class="h5 mb-3">Popis članova</h2>
  <div class="table-responsive">
    <table class="table table-sm align-middle">
      <thead>
        <tr>
          <th>ID</th><th>Ime i prezime</th><th>Grupa</th><th>Aktivan</th><th>Članarina (€)</th><th>Liječnička</th><th>Dokumenti</th><th>Akcije</th>
        </tr>
      </thead>
      <tbody>
        {% for m in members %}
          {% set warn = (m.lijecnicka_dana_preostalo is not none and m.lijecnicka_dana_preostalo <= 14) %}
          <tr class="{{ 'table-danger' if warn }}">
            <td>{{ m.id }}</td>
            <td>
              <div class="fw-semibold">{{ m.ime_prezime }}</div>
              <div class="text-muted small">{{ m.datum_rodjenja }} · OIB: {{ m.oib }}</div>
            </td>
            <td>{{ m.grupa or '—' }}</td>
            <td>{{ 'Da' if m.aktivan_natjecatelj else 'Ne' }}</td>
            <td>{{ (m.iznos_clanarine or 0)|round(2) if m.placa_clanarinu else 'Ne plaća' }}</td>
            <td>
              {% if m.lijecnicka_vrijedi_do %}
                {{ m.lijecnicka_vrijedi_do }}
                {% if warn %}<span class="danger-badge ms-2">Istječe za {{ m.lijecnicka_dana_preostalo }} dana</span>{% endif %}
              {% else %} — {% endif %}
            </td>
            <td class="small">
              {% if m.pristupnica_pdf %}<a href="{{ m.pristupnica_pdf }}" target="_blank">Pristupnica (PDF)</a><br>{% endif %}
              {% if m.privola_pdf %}<a href="{{ m.privola_pdf }}" target="_blank">Privola (PDF)</a><br>{% endif %}
              {% if m.pristupnica_upload %}<a href="{{ m.pristupnica_upload }}" target="_blank">Pristupnica (upl.)</a><br>{% endif %}
              {% if m.privola_upload %}<a href="{{ m.privola_upload }}" target="_blank">Privola (upl.)</a><br>{% endif %}
              {% if m.lijecnicka_upload %}<a href="{{ m.lijecnicka_upload }}" target="_blank">Liječnička (upl.)</a>{% endif %}
            </td>
            <td>
              <div class="d-flex flex-wrap gap-2">
                <button class="btn btn-sm btn-outline-secondary" onclick='return fillForm({{ m|tojson }});'>Uredi</button>
                <form method="POST" action="{{ url_for('clanovi_obrisi', mid=m.id) }}" onsubmit="return confirm('Obrisati člana?');">
                  <button class="btn btn-sm btn-outline-danger">Obriši</button>
                </form>
                {% if m.pristupnica_pdf %}<a class="btn btn-sm btn-outline-accent" href="{{ m.pristupnica_pdf }}" target="_blank">Pristupnica PDF</a>{% endif %}
                {% if m.privola_pdf %}<a class="btn btn-sm btn-outline-accent" href="{{ m.privola_pdf }}" target="_blank">Privola PDF</a>{% endif %}
              </div>
            </td>
          </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>

<script>
  function fillForm(m){
    document.getElementById('member_id').value = m.id;
    document.getElementById('f_ime').value = m.ime_prezime||"";
    document.getElementById('f_dat').value = m.datum_rodjenja||"";
    document.getElementById('f_spol').value = m.spol||"";
    document.getElementById('f_oib').value = m.oib||"";
    document.getElementById('f_mjesto').value = m.mjesto||"";
    document.getElementById('f_em1').value = m.email_sportas||"";
    document.getElementById('f_em2').value = m.email_roditelj||"";
    document.getElementById('f_obr').value = m.osobna_broj||"";
    document.getElementById('f_ovd').value = m.osobna_vrijedi_do||"";
    document.getElementById('f_oiz').value = m.osobna_izdao||"";
    document.getElementById('f_pbr').value = m.putovnica_broj||"";
    document.getElementById('f_pvd').value = m.putovnica_vrijedi_do||"";
    document.getElementById('f_piz').value = m.putovnica_izdao||"";
    document.getElementById('f_grp').value = m.grupa||"";
    document.getElementById('f_act').checked = !!m.aktivan_natjecatelj;
    document.getElementById('f_vet').checked = !!m.veteran;
    document.getElementById('f_ost').checked = !!m.ostalo;
    document.getElementById('f_fee').checked = !!m.placa_clanarinu;
    document.getElementById('f_amt').value = (m.iznos_clanarine ?? 30);
    return false;
  }
</script>
""" + TPL_BASE_FOOT

# =========================================================
# Pokretanje (NE DIRATI)
# =========================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

